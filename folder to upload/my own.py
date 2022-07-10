import cv2
import numpy as np
import utilis
from openpyxl import Workbook, load_workbook
import traceback
###################################
widthImg=700
heightImg=1200
questions = 75
questionsperC=25
choices=4
widthCont=700
heightCont=1200
threshold_pixels=3800
mindif=500#how much more than average pixels
with open("answer key.txt") as a:
   answer_key= a.readlines()

#################################3
webcamFeed = True
cameraNo=1
cap = cv2.VideoCapture(cameraNo)
############importing student list##########
workbook=load_workbook("student list.xlsx")
worksheet= workbook.active
name_list=[]

for n in range(1,worksheet.max_row+1):

    name_list.append(worksheet.cell(n,2).value)

print(name_list)

#

#CREATING EXCEL
wb= load_workbook("report.xlsx")

while True:
    ret, img = cap.read()
    img = cv2.rotate(img,cv2.ROTATE_90_CLOCKWISE)

    cv2.imshow('feed',img)
    ##############################LOAD PICTURE#################
    if cv2.waitKey(1) & 0xFF == ord('s'):
        #img=cv2.imread("pictures\last.jfif")
        #preprocessing
        img=cv2.resize(img,(widthImg,heightImg))
        imgContours = img.copy()
        imgBiggestContours = img.copy()
        imgGray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        imgBlur= cv2.GaussianBlur(imgGray,(5,5),1)
        imgCanny = cv2.Canny(imgBlur,10,50)
        try:

            #finding all contours
            contours, heirarchy = cv2.findContours(imgCanny,cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
            cv2.drawContours(imgContours,contours,-1,(0,255,0),10)
            #find rectangles
            rectCon=utilis.rectContour(contours)
            firstContoure = utilis.getCornerPoints(rectCon[0])#the 1 to 25 contour, change index to get other contours
            secondContoure = utilis.getCornerPoints(rectCon[1])
            thirdContoure = utilis.getCornerPoints(rectCon[2])#the 1 to 25 contour, change index to get other contours
            rollContour = utilis.getCornerPoints(rectCon[3])#the 1 to 25 contour, change index to get other contours# #the 1 to 25 contour, change index to get other contours

            def funk(sim):
                try:
                    sim= int(str(sim).split(" ")[0].replace("[[", ""))
                    return sim
                except:
                    return 0

            list=[funk(firstContoure[0]), funk(secondContoure[0]), funk(thirdContoure[0])]
            list.sort()

            if list[0]==funk(firstContoure[0]):
                firstContour=firstContoure
                if list[1]== funk(secondContoure[0]):
                    secondContour=secondContoure
                    thirdContour=thirdContoure
                elif list[1]==funk(thirdContoure[0]):
                    secondContour = thirdContoure
                    thirdContour= secondContoure
            elif list[0]==funk(secondContoure[0]):
                firstContour=secondContoure
                if list[1]==funk(firstContoure[0]):
                    secondContour= firstContoure
                    thirdContour = thirdContoure
                elif list[1]== funk(thirdContoure[0]):
                    secondContour= thirdContoure
                    thirdContour=firstContoure

            elif list[0]==funk(thirdContoure[0]):
                firstContour=thirdContoure
                if list[1]==funk(firstContoure[0]):
                    secondContour= firstContoure
                    thirdContour = secondContoure
                elif list[1]== funk(secondContoure[0]):
                    secondContour= secondContoure
                    thirdContour=firstContoure




            #if biggestContour.size!=0 and gradePoints != 0:
            cv2.drawContours(imgBiggestContours,firstContour,-1,(0,255,0),20)
            cv2.drawContours(imgBiggestContours, secondContour, -1, (255, 0, 0), 20)
            cv2.drawContours(imgBiggestContours, thirdContour, -1, (0, 0, 255), 20)
            cv2.drawContours(imgBiggestContours, rollContour, -1, (0, 255, 255), 20)

            firstContour=utilis.reorder(firstContour)
            secondContour=utilis.reorder(secondContour)
            thirdContour=utilis.reorder(thirdContour)
            rollContour=utilis.reorder(rollContour)


            ######################################ROLL CONTUR#################
            pt1 = np.float32(rollContour)
            rollwidthImg=700
            rollheightImg=501
            pt2= np.float32([[0,0],[rollwidthImg,0],[0,rollheightImg],[rollwidthImg,rollheightImg]])
            matrix=cv2.getPerspectiveTransform(pt1,pt2)
            imgWarpColored = cv2.warpPerspective(img,matrix,(rollwidthImg,rollheightImg))




            imgWarpGray=cv2.cvtColor(imgWarpColored,cv2.COLOR_BGR2GRAY)
            imgThresh= cv2.threshold(imgWarpGray,150,255,cv2.THRESH_BINARY_INV)[1]
            #cv2.imshow("kjb", imgWarpColored)

            boxes =utilis.splitBoxes(imgThresh,3,10)


            #getting NON ZERO pixel values of each BOX
            myPixelVal=np.zeros((3,10))
            countC=0
            countR=0

            for image in boxes:
                totalPixels = cv2.countNonZero(image)
                myPixelVal[countR][countC]=totalPixels
                countC+=1
                if (countC== 10):
                    countR+=1
                    countC=0
            #print(myPixelVal)
            #threshold = 5900
            roll=''
            for x in range(0,3):
                arr = myPixelVal[x]
                myIndexVal = np.where(arr==np.amax(arr))
                roll = roll+(str(myIndexVal[0]).replace("[",'').replace(']',''))

            roll =roll.replace('0','')
            try:
                ws = wb["roll-"+roll]
            except:
                print(roll)



            #####################################PROCESESSING THREE CONTOURS
            #                                        cONTOUR 1#######################################
            pt1 = np.float32(firstContour)
            pt2= np.float32([[0,0],[widthImg,0],[0,heightImg],[widthImg,heightImg]])
            matrix=cv2.getPerspectiveTransform(pt1,pt2)
            imgWarpColored = cv2.warpPerspective(img,matrix,(widthImg,heightImg))

            #ptG1 = np.float32(gradePoints)
            #ptG2= np.float32([[0,0],[325,0],[0,150],[325,150]])
            #matrixG=cv2.getPerspectiveTransform(ptG1,ptG2)
            #imgGradeDisplay = cv2.warpPerspective(img,matrixG,(325,150))
            #APPLY THRESHOLD to first countour
            imgWarpGray=cv2.cvtColor(imgWarpColored,cv2.COLOR_BGR2GRAY)
            imgThresh= cv2.threshold(imgWarpGray,150,255,cv2.THRESH_BINARY_INV)[1]

            boxes =utilis.splitBoxes(imgThresh,questionsperC,choices)

            CONT1=imgThresh
            #getting NON ZERO pixel values of each BOX
            myPixelVal=np.zeros((questionsperC,choices))
            countC=0
            countR=0

            for image in boxes:
                totalPixels = cv2.countNonZero(image)
                myPixelVal[countR][countC]=totalPixels
                countC+=1
                if (countC== choices):
                    countR+=1
                    countC=0
            #print(myPixelVal)

            ##############################COMMENT OUT FROM HERE TO GET THE THRESHOLD PIXEL VALUES


            for x in range (0,questionsperC):
                arr=myPixelVal[x]
                average = np.average(arr)

                myIndexVal = np.where(arr > threshold_pixels)
                list = myIndexVal[0]
                value=""
                for stanswer in list:
                        if stanswer == 0:
                            value=value+"A"

                        if stanswer == 1:
                            value = value + "B"

                        if stanswer == 2:
                            value = value + "C"

                        if stanswer == 3:
                            value = value + "D"
                ws.cell(column=2, row=6 + int(x), value=value)
                #print(myIndexVal[0])
                #myIndex.append(myIndexVal[0][0])


            ########################################################################End of countour 1####################


            ##############################################################CONTOUR 2#################################

            pt1 = np.float32(secondContour)
            pt2= np.float32([[0,0],[widthImg,0],[0,heightImg],[widthImg,heightImg]])
            matrix=cv2.getPerspectiveTransform(pt1,pt2)
            imgWarpColored = cv2.warpPerspective(img,matrix,(widthImg,heightImg))



            imgWarpGray=cv2.cvtColor(imgWarpColored,cv2.COLOR_BGR2GRAY)
            imgThresh= cv2.threshold(imgWarpGray,150,255,cv2.THRESH_BINARY_INV)[1]
            CONT2=imgThresh
            boxes =utilis.splitBoxes(imgThresh,questionsperC,choices)


            #getting NON ZERO pixel values of each BOX
            myPixelVal=np.zeros((questionsperC,choices))
            countC=0
            countR=0

            for image in boxes:
                totalPixels = cv2.countNonZero(image)
                myPixelVal[countR][countC]=totalPixels

                countC+=1
                if (countC== choices):
                    countR+=1
                    countC=0
            #print(myPixelVal)
            #FINDING INDEX VALUES OF THE MARKINGS
            myIndex=[]
            ##############################COMMENT OUT FROM HERE TO GET THE THRESHOLD PIXEL VALUES


            for x in range (0,questionsperC):
                arr=myPixelVal[x]
                average=np.average(arr)

                #= np.where(arr==np.amax(arr))

                myIndexVal = np.where(arr>threshold_pixels)

                list = myIndexVal[0]
                value=""
                for stanswer in list:
                        if stanswer == 0:
                            value=value+"A"

                        if stanswer == 1:
                            value = value + "B"

                        if stanswer == 2:
                            value = value + "C"

                        if stanswer == 3:
                            value = value + "D"
                ws.cell(column=2, row=31 + int(x), value=value)
            #######################################################END OF CONTOUR 2###########################################


            #######################################################CONTOUR 3#############################################
            pt1 = np.float32(thirdContour)
            pt2= np.float32([[0,0],[widthImg,0],[0,heightImg],[widthImg,heightImg]])
            matrix=cv2.getPerspectiveTransform(pt1,pt2)
            imgWarpColored = cv2.warpPerspective(img,matrix,(widthImg,heightImg))



            imgWarpGray=cv2.cvtColor(imgWarpColored,cv2.COLOR_BGR2GRAY)
            imgThresh= cv2.threshold(imgWarpGray,150,255,cv2.THRESH_BINARY_INV)[1]
            CONT3=imgThresh
            boxes =utilis.splitBoxes(imgThresh,questionsperC,choices)
            #cv2.imshow("box",boxes[99])

            #getting NON ZERO pixel values of each BOX
            myPixelVal=np.zeros((questionsperC,choices))
            countC=0
            countR=0

            for image in boxes:
                totalPixels = cv2.countNonZero(image)
                myPixelVal[countR][countC]=totalPixels
                countC+=1
                if (countC== choices):
                    countR+=1
                    countC=0
            #print(myPixelVal)
            #FINDING INDEX VALUES OF THE MARKINGS
            myIndex=[]
            ##############################COMMENT OUT FROM HERE TO GET THE THRESHOLD PIXEL VALUES


            for x in range (0,questionsperC):
                arr=myPixelVal[x]
                average = np.average(arr)

                myIndexVal = np.where(arr > threshold_pixels)
                list = myIndexVal[0]
                value=""
                for stanswer in list:
                        if stanswer == 0:
                            value=value+"A"

                        if stanswer == 1:
                            value = value + "B"

                        if stanswer == 2:
                            value = value + "C"

                        if stanswer == 3:
                            value = value + "D"
                ws.cell(column=2, row=56 + int(x), value=value)
            #######################################################END OF CONTOUR 3###########################################


            #GRADING
            #grading
            for row in range(6,6+questions):

                stanswer=str(ws.cell(column=2, row =row).value).strip()
                canswer=str(ws.cell(column=3, row =row).value).strip()


                if stanswer == canswer:

                    ws.cell(column=4, row=row, value= 4)
                elif stanswer == '' or stanswer is None:
                    ws.cell(column=4, row=row, value=0)
                else:
                    ws.cell(column=4, row=row, value=-1)


            #totalling marks
            ws.cell(column=4, row=1, value='Total marks')
            sum=0
            for row in range(6,6+len(answer_key)):
                sum=sum + ws.cell(column=4, row=row).value


            ws.cell(column=5, row=1, value=sum)




            wb.save("report.xlsx")
            imgBlank = np.zeros_like(img)
            imageArray = ([imgContours, imgBiggestContours, CONT1, CONT2, CONT3])

            imgStacked = utilis.stackImages(imageArray, 0.5)
            imgBiggestContours = cv2.resize(imgBiggestContours, (500, 900))
            cv2.imshow("Original", imgStacked)
            if cv2.waitKey(1) & 0xFF == ord('s'):
                cv2.destroyAllWindows

        except:
            print("image not good enough")
            traceback.print_exc()

            cv2.imshow("Original", imgContours)

    if cv2.waitKey(1) & 0xFF == ord('1'):
        break
        cv2.destroyAllWindows

