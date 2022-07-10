from openpyxl import Workbook, load_workbook
with open("answer key.txt") as a:
   answer_key= a.readlines()
wb= Workbook()
workbook=load_workbook("student list.xlsx")
worksheet= workbook.active
name_list=[]

for n in range(1,worksheet.max_row+1):

    name_list.append(worksheet.cell(n,2).value)

for x in range(len(name_list)):
    wb.create_sheet('roll-'+str(x+1))

    ws = wb['roll-'+str(x+1)]
    ws['A1'] = "Name:"
    ws['B1'] = name_list[x]

    ws['A2'] = "Roll:"
    ws['B2'] = str(x+1)
    ws['A3'] = "Batch:"
    ws['A5'] = "Sr."
    ws['B5'] = "Your Answer"
    ws['C5'] = "Correct Answer"
    ws['D5'] = "Marks awarded"
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    for answer_pairs in answer_key:
        q_number = answer_pairs.split(':')[0]
        ws.cell(column=1, row=5 + int(q_number), value=q_number)
        canswer = answer_pairs.split(':')[1].strip()
        ws.cell(column=3, row=5 + int(q_number), value=canswer)

wb.save("report.xlsx")